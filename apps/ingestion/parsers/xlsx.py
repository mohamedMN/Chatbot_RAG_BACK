# backend/apps/ingestion/parsers/xlsx.py
from typing import List
from io import BytesIO
from openpyxl import load_workbook
from apps.ingestion.schema import Section


def parse_xlsx_bytes(data: bytes, filename: str, max_cells: int = 50000) -> List[Section]:
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    out: List[Section] = []
    seen = 0
    ordinal = 0
    for ws in wb.worksheets:
        lines = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).strip()
                if s:
                    lines.append(s)
                seen += 1
                if seen >= max_cells:
                    break
            if seen >= max_cells:
                break
        txt = "\n".join(lines).strip()
        if txt:
            out.append(Section(subject=ws.title[:120] or f"Feuille {ws.title}",
                               source=f"{filename}#sheet={ws.title}",
                               content=txt, ordinal=ordinal))
            ordinal += 1
        if seen >= max_cells:
            break
    return out
